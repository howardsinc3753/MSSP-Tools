#!/usr/bin/env python3
"""
USE CASE 4: Daily Consumption Data Pull (Billing & Reporting)

PURPOSE:
    Pull daily point consumption for all customers to generate monthly invoices,
    export to Excel/CSV for billing teams, and optionally store in PostgreSQL
    for historical analysis beyond the 3-month portal retention.

WHAT IT DOES:
    1. Retrieves consumption data from FortiFlex API for specified date range
    2. Displays formatted console report with daily breakdowns
    3. Exports to JSON (default), Excel (.xlsx), or CSV formats
    4. Optionally saves to PostgreSQL database for long-term storage
    5. Supports filtering by serial number or pulling all entitlements

SAFE TO RUN: YES (read-only operation)

REQUIREMENTS:
    - credentials.json configured with account_id
    - Valid FortiFlex MSSP program
    - For Excel export: pip install openpyxl pandas
    - For database: pip install psycopg2-binary + PostgreSQL setup

USAGE:
    # Console report only (default)
    python consumption_report_v2.py --days 7

    # Export to Excel workbook
    python consumption_report_v2.py --days 30 --output excel

    # Export to CSV
    python consumption_report_v2.py --days 7 --output csv

    # Save to database (requires PostgreSQL setup)
    python consumption_report_v2.py --days 1 --save-to-db

    # Specific device with Excel export
    python consumption_report_v2.py --serial FMVMMLTM12345 --days 7 --output excel

See EXAMPLES_SUMMARY.md for complete documentation and sample outputs.
"""

import sys
import os
import json
import argparse
from datetime import datetime, timedelta

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from fortiflex_client import FortiFlexClient, get_oauth_token

# Optional imports for export features
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import psycopg2
    from psycopg2.extras import execute_values
    DATABASE_AVAILABLE = True
except ImportError:
    DATABASE_AVAILABLE = False


def load_credentials():
    """Load credentials from config file."""
    config_file = os.path.join(os.path.dirname(__file__), '..', 'testing', 'config', 'credentials.json')

    if not os.path.exists(config_file):
        print("[ERROR] credentials.json not found!")
        print(f"   Expected location: {config_file}")
        print("\n   Please run: python testing/discover_program.py")
        sys.exit(1)

    with open(config_file, 'r') as f:
        creds = json.load(f)

    # Validate required fields
    required_fields = ['api_username', 'api_password', 'program_serial_number', 'account_id']
    for field in required_fields:
        if field not in creds.get('fortiflex', {}):
            print(f"[ERROR] Missing required field '{field}' in credentials.json")
            print(f"\n   Your credentials.json must include:")
            print(f"   {{\n     \"fortiflex\": {{")
            print(f"       \"api_username\": \"YOUR_USERNAME\",")
            print(f"       \"api_password\": \"YOUR_PASSWORD\",")
            print(f"       \"program_serial_number\": \"ELAVMSXXXXXXXX\",")
            print(f"       \"account_id\": YOUR_ACCOUNT_ID")
            print(f"     }}\n   }}")
            sys.exit(1)

    return creds


def list_all_configurations(client: FortiFlexClient):
    """
    List all configurations to show what entitlements exist.

    Args:
        client: FortiFlexClient instance

    Returns:
        List of configurations
    """
    print(f"\n{'='*80}")
    print("AVAILABLE CONFIGURATIONS (Templates)")
    print(f"{'='*80}\n")

    configs = client.list_configs()

    if not configs or 'configs' not in configs:
        print("No configurations found.")
        return []

    config_list = configs['configs']

    print(f"{'ID':<8} {'Name':<30} {'Product Type':<25} {'Status':<10}")
    print("-" * 80)

    for cfg in config_list:
        cfg_id = cfg.get('id', 'N/A')
        name = cfg.get('name', 'Unknown')
        product = cfg.get('productType', {}).get('name', 'Unknown')
        status = cfg.get('status', 'Unknown')

        print(f"{cfg_id:<8} {name:<30} {product:<25} {status:<10}")

    print(f"\n{'='*80}\n")
    print(f"[NOTE] These are configuration templates, not actual entitlements.")
    print(f"       To see actual deployed entitlements, you need to query the portal")
    print(f"       or use the consumption API (which only shows devices consuming points).\n")

    return config_list


def get_consumption_report(client: FortiFlexClient, account_id: int, days_back: int = 30, serial_filter: str = None):
    """
    Get consumption report for specified number of days.

    Args:
        client: FortiFlexClient instance
        account_id: FortiFlex account ID (required per API spec)
        days_back: Number of days to look back (default: 30)
        serial_filter: Optional serial number to filter by

    Returns:
        Consumption data dict
    """
    print(f"\n{'='*80}")
    if serial_filter:
        print(f"CONSUMPTION REPORT - {serial_filter} - LAST {days_back} DAYS")
    else:
        print(f"CONSUMPTION REPORT - ALL ENTITLEMENTS - LAST {days_back} DAYS")
    print(f"{'='*80}\n")

    # Calculate date range - use local date (not datetime) to match GUI
    today = datetime.now().date()
    end_date = today  # inclusive end as YYYY-MM-DD
    start_date = today - timedelta(days=days_back)

    start_str = start_date.strftime("%Y-%m-%d")
    end_str = end_date.strftime("%Y-%m-%d")

    print(f"Date Range: {start_str} to {end_str}")
    print(f"Account ID: {account_id}")
    if serial_filter:
        print(f"Serial Number Filter: {serial_filter}")
    print(f"\nRetrieving consumption data...\n")

    # Get consumption data with required identifiers
    try:
        consumption = client.get_entitlement_points(
            account_id=account_id,  # REQUIRED per API spec
            serial_number=serial_filter,  # Optional filter
            start_date=start_str,
            end_date=end_str
        )
    except Exception as e:
        msg = str(e)
        if "400" in msg:
            print(f"\n[INFO] No consumption data found for the date range.")
            print(f"       (HTTP 400) Server message:")
            print(f"       {msg[:500]}")
            print("\n       Possible reasons:")
            print("       1. Entitlements exist but haven't consumed points yet")
            print("       2. Entitlements are in PENDING status (not activated)")
            print("       3. Devices haven't checked in with FortiCloud")
            print("       4. No entitlements exist for this date range")
            print("\n[NOTE] The '/entitlements/points' API only returns consumption data.")
            print("       It does NOT return a list of entitlements.")
            print("\n       Once ACTIVE devices check in with FortiCloud and start")
            print("       consuming points, they will appear in this report.\n")
            return None
        else:
            raise

    # Check if we got data
    points_data = consumption.get('entitlements', [])

    if not points_data:
        print(f"[INFO] API 200 OK but no consumption for {start_str} to {end_str}.")
        print("       This means entitlements exist but have not consumed any points yet.")
        print("\n       Your entitlements are configured, but they need to:")
        print("       1. Be activated (if PENDING)")
        print("       2. Check in with FortiCloud")
        print("       3. Start consuming services\n")
        return consumption

    # Optional defensive client-side filter (server already filtered if serialNumber was sent)
    if serial_filter:
        points_data = [p for p in points_data if p.get('serialNumber') == serial_filter]

        if not points_data:
            print(f"[INFO] No consumption data found for serial number: {serial_filter}")
            print(f"       Either this device hasn't consumed points, or the serial number is incorrect.\n")
            return None

    # Process the data
    print(f"[SUCCESS] Retrieved consumption data for {len(points_data)} entitlement(s)\n")

    # Display summary
    print(f"{'='*80}")
    print("CONSUMPTION SUMMARY")
    print(f"{'='*80}\n")

    # Sort by total points (highest first)
    sorted_points = sorted(points_data, key=lambda x: x.get('points', 0), reverse=True)

    # Display header
    print(f"{'Serial Number':<30} {'Account ID':<15} {'Total Points':>15}")
    print("-" * 65)

    # Display each entitlement
    grand_total = 0
    for item in sorted_points:
        serial = item.get('serialNumber', 'Unknown')
        account_id = item.get('accountId', 'N/A')
        points = item.get('points', 0)
        grand_total += points

        print(f"{serial:<30} {str(account_id):<15} {points:>15.2f}")

    # Display totals
    print("-" * 65)
    print(f"{'GRAND TOTAL':<30} {'':<15} {grand_total:>15.2f}")
    print(f"\n{'='*80}\n")

    # Daily breakdown - be tolerant to shape differences
    if sorted_points and isinstance(sorted_points[0].get('daily'), list):
        for idx, item in enumerate(sorted_points, 1):
            serial = item.get('serialNumber', 'Unknown')
            daily_data = item.get('daily', [])

            if not daily_data:
                continue

            print(f"{'='*80}")
            print(f"DAILY BREAKDOWN #{idx} - {serial}")
            print(f"{'='*80}\n")

            print(f"{'Date':<15} {'Points':>15}")
            print("-" * 35)

            for day in daily_data:
                date_str = day.get('date', 'Unknown')
                points = day.get('points', 0)
                print(f"{date_str:<15} {points:>15.2f}")

            print(f"\n{'='*80}\n")

    return consumption


def export_to_excel(consumption_data: dict, filename: str, days_back: int):
    """
    Export consumption data to Excel workbook with formatting.

    Args:
        consumption_data: Consumption data from API
        filename: Output filename (without extension)
        days_back: Number of days in report

    Returns:
        Path to created Excel file
    """
    if not EXCEL_AVAILABLE:
        print("[ERROR] Excel export requires: pip install openpyxl pandas")
        return None

    points_data = consumption_data.get('entitlements', [])
    if not points_data:
        print("[INFO] No data to export")
        return None

    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    headers = ["Serial Number", "Account ID", "Total Points", "Days"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Write data
    sorted_points = sorted(points_data, key=lambda x: x.get('points', 0), reverse=True)
    grand_total = 0

    for row, item in enumerate(sorted_points, 2):
        ws_summary.cell(row=row, column=1, value=item.get('serialNumber', 'Unknown')).border = border
        ws_summary.cell(row=row, column=2, value=item.get('accountId', 'N/A')).border = border
        points = item.get('points', 0)
        ws_summary.cell(row=row, column=3, value=points).border = border
        ws_summary.cell(row=row, column=4, value=days_back).border = border
        grand_total += points

    # Add total row
    total_row = len(sorted_points) + 2
    ws_summary.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
    ws_summary.cell(row=total_row, column=3, value=grand_total).font = Font(bold=True)

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 10

    # Sheet 2: Daily Details
    if sorted_points and isinstance(sorted_points[0].get('daily'), list):
        ws_daily = wb.create_sheet("Daily Details")

        # Write headers
        daily_headers = ["Serial Number", "Date", "Points", "Account ID"]
        for col, header in enumerate(daily_headers, 1):
            cell = ws_daily.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Write daily data
        row = 2
        for item in sorted_points:
            serial = item.get('serialNumber', 'Unknown')
            account_id = item.get('accountId', 'N/A')
            daily_data = item.get('daily', [])

            for day in daily_data:
                ws_daily.cell(row=row, column=1, value=serial).border = border
                ws_daily.cell(row=row, column=2, value=day.get('date', 'Unknown')).border = border
                ws_daily.cell(row=row, column=3, value=day.get('points', 0)).border = border
                ws_daily.cell(row=row, column=4, value=account_id).border = border
                row += 1

        # Adjust column widths
        ws_daily.column_dimensions['A'].width = 30
        ws_daily.column_dimensions['B'].width = 15
        ws_daily.column_dimensions['C'].width = 12
        ws_daily.column_dimensions['D'].width = 15

    # Save workbook
    excel_file = f"{filename}.xlsx"
    wb.save(excel_file)
    print(f"[SUCCESS] Excel report saved to: {excel_file}")
    return excel_file


def export_to_csv(consumption_data: dict, filename: str):
    """
    Export consumption data to CSV file.

    Args:
        consumption_data: Consumption data from API
        filename: Output filename (without extension)

    Returns:
        Path to created CSV file
    """
    if not EXCEL_AVAILABLE:  # pandas is needed for CSV too
        print("[ERROR] CSV export requires: pip install pandas")
        return None

    points_data = consumption_data.get('entitlements', [])
    if not points_data:
        print("[INFO] No data to export")
        return None

    # Create DataFrame
    rows = []
    for item in points_data:
        serial = item.get('serialNumber', 'Unknown')
        account_id = item.get('accountId', 'N/A')
        total_points = item.get('points', 0)

        # Add daily breakdown if available
        daily_data = item.get('daily', [])
        if daily_data:
            for day in daily_data:
                rows.append({
                    'Serial Number': serial,
                    'Account ID': account_id,
                    'Date': day.get('date', 'Unknown'),
                    'Points': day.get('points', 0),
                    'Total Points': total_points
                })
        else:
            rows.append({
                'Serial Number': serial,
                'Account ID': account_id,
                'Date': 'N/A',
                'Points': total_points,
                'Total Points': total_points
            })

    df = pd.DataFrame(rows)
    csv_file = f"{filename}.csv"
    df.to_csv(csv_file, index=False)
    print(f"[SUCCESS] CSV report saved to: {csv_file}")
    return csv_file


def save_to_database(consumption_data: dict, db_config: dict):
    """
    Save consumption data to PostgreSQL database.

    Args:
        consumption_data: Consumption data from API
        db_config: Database connection configuration

    Returns:
        Number of records inserted
    """
    if not DATABASE_AVAILABLE:
        print("[ERROR] Database support requires: pip install psycopg2-binary")
        return 0

    points_data = consumption_data.get('entitlements', [])
    if not points_data:
        print("[INFO] No data to save")
        return 0

    try:
        # Connect to database
        conn = psycopg2.connect(
            host=db_config.get('host', 'localhost'),
            port=db_config.get('port', 5432),
            database=db_config.get('database', 'fortiflex'),
            user=db_config.get('username', 'app'),
            password=db_config.get('password', '')
        )
        cursor = conn.cursor()

        # Prepare data for insertion
        records = []
        for item in points_data:
            serial = item.get('serialNumber')
            account_id = item.get('accountId')
            config_id = item.get('configId')

            daily_data = item.get('daily', [])
            for day in daily_data:
                records.append((
                    serial,
                    account_id,
                    config_id,
                    day.get('date'),
                    day.get('points', 0),
                    datetime.now()
                ))

        # Insert data (using ON CONFLICT to avoid duplicates)
        insert_query = """
            INSERT INTO consumption_daily (serial_number, account_id, config_id, date, points, recorded_at)
            VALUES %s
            ON CONFLICT (serial_number, date) DO UPDATE
            SET points = EXCLUDED.points,
                account_id = EXCLUDED.account_id,
                config_id = EXCLUDED.config_id,
                recorded_at = EXCLUDED.recorded_at
        """

        execute_values(cursor, insert_query, records)
        conn.commit()

        print(f"[SUCCESS] Saved {len(records)} records to database")

        cursor.close()
        conn.close()

        return len(records)

    except Exception as e:
        print(f"[ERROR] Database save failed: {str(e)}")
        print("[INFO] Make sure PostgreSQL is running and database schema is created")
        print("       See database/schema.sql for table definitions")
        return 0


def main():
    """Main execution."""

    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description='FortiFlex Consumption Report',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # All entitlements, last 30 days (default)
  python consumption_report_v2.py

  # Daily report (last 1 day)
  python consumption_report_v2.py --days 1

  # Specific entitlement, last 30 days
  python consumption_report_v2.py --serial FMVMMLTMXXXXXXXX

  # Specific entitlement, last 7 days
  python consumption_report_v2.py --serial FMVMMLTMXXXXXXXX --days 7

  # Show configurations first
  python consumption_report_v2.py --list-configs
        """
    )

    parser.add_argument(
        '--days',
        type=int,
        default=30,
        help='Number of days to look back (default: 30)'
    )

    parser.add_argument(
        '--serial',
        type=str,
        help='Filter by specific serial number (e.g., FMVMMLTMXXXXXXXX)'
    )

    parser.add_argument(
        '--list-configs',
        action='store_true',
        help='List all configurations before running report'
    )

    parser.add_argument(
        '--output',
        type=str,
        choices=['json', 'excel', 'csv'],
        default='json',
        help='Output format: json (default), excel, or csv'
    )

    parser.add_argument(
        '--save-to-db',
        action='store_true',
        help='Save consumption data to PostgreSQL database (requires database setup)'
    )

    args = parser.parse_args()

    # Load credentials
    creds = load_credentials()
    API_USERNAME = creds['fortiflex']['api_username']
    API_PASSWORD = creds['fortiflex']['api_password']
    PROGRAM_SN = creds['fortiflex']['program_serial_number']
    ACCOUNT_ID = creds['fortiflex']['account_id']  # REQUIRED per API spec

    try:
        print(f"\n{'='*80}")
        print("FORTIFLEX CONSUMPTION REPORT")
        print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*80}\n")

        # Authenticate
        print("Authenticating with FortiFlex API...")
        token = get_oauth_token(API_USERNAME, API_PASSWORD, client_id="flexvm")
        print("[SUCCESS] Authentication successful\n")

        # Initialize client
        client = FortiFlexClient(token, PROGRAM_SN)
        print(f"Program Serial Number: {PROGRAM_SN}")
        print(f"Account ID: {ACCOUNT_ID}\n")

        # List configurations if requested
        if args.list_configs:
            list_all_configurations(client)

        # Get consumption report
        consumption = get_consumption_report(
            client,
            account_id=ACCOUNT_ID,
            days_back=args.days,
            serial_filter=args.serial
        )

        if consumption is None:
            print("No data to report.")
            return 1

        # Generate output filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_serial = (args.serial or "all").replace(":", "_").replace("/", "_")
        base_filename = f"consumption_{safe_serial}_{args.days}days_{timestamp}"

        # Export based on format
        if args.output == 'excel':
            export_to_excel(consumption, base_filename, args.days)
        elif args.output == 'csv':
            export_to_csv(consumption, base_filename)
        else:  # json (default)
            output_file = f"{base_filename}.json"
            with open(output_file, 'w') as f:
                json.dump({
                    'generated_at': datetime.now().isoformat(),
                    'days_back': args.days,
                    'serial_filter': args.serial,
                    'program_serial': PROGRAM_SN,
                    'consumption': consumption
                }, f, indent=2)
            print(f"[SUCCESS] JSON report saved to: {output_file}\n")

        # Save to database if requested
        if args.save_to_db:
            print("\n[INFO] Saving to database...")
            db_config = creds.get('database', {})
            if not db_config:
                print("[WARNING] No database configuration found in credentials.json")
                print("         Add 'database' section with host, port, database, username, password")
            else:
                save_to_database(consumption, db_config)

        print(f"{'='*80}")
        print(f"Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*80}\n")

        return 0

    except Exception as e:
        print(f"\n[ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
